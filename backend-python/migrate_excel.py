"""
Migration script: Reads Bodegas Sinaloa_bajio.xlsx and populates the database.
- Drops and recreates regiones + bodegas tables
- Creates regiones from unique DDR per state
- Inserts 1048 bodegas with all fields from the Excel
"""
import os
import re
import sys
import psycopg2
from psycopg2.extras import RealDictCursor
from dotenv import load_dotenv

load_dotenv()

DB_HOST = os.getenv("DB_HOST", "localhost")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_USER = os.getenv("DB_USER", "jesus")
DB_PASSWORD = os.getenv("DB_PASSWORD", "2025")
DB_NAME = os.getenv("DB_NAME", "bodegas")

# Try local Excel first, then parent directory
EXCEL_PATH = None
for p in [
    os.path.join(os.path.dirname(__file__), "Bodegas Sinaloa_bajio.xlsx"),
    os.path.join(os.path.dirname(__file__), "..", "Bodegas Sinaloa_bajio.xlsx"),
]:
    if os.path.exists(p):
        EXCEL_PATH = p
        break

if not EXCEL_PATH:
    print("ERROR: Excel file not found")
    sys.exit(1)

print(f"Using Excel: {EXCEL_PATH}")


def parse_storage(val):
    """Parse storage value like '4,000' or '52,600' to float."""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace(",", "").replace(" ", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def main():
    import openpyxl

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    conn = psycopg2.connect(
        host=DB_HOST, port=DB_PORT, user=DB_USER, password=DB_PASSWORD, database=DB_NAME
    )
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # --- Schema migration ---
    print("Migrating schema...")
    cur.execute("""
        DROP TABLE IF EXISTS bodegas CASCADE;
        DROP TABLE IF EXISTS regiones CASCADE;

        CREATE TABLE regiones (
            id SERIAL PRIMARY KEY,
            nombre VARCHAR(100) NOT NULL,
            estado VARCHAR(100),
            UNIQUE(nombre, estado)
        );

        CREATE TABLE bodegas (
            id SERIAL PRIMARY KEY,
            clave VARCHAR(50),
            nombre VARCHAR(500) NOT NULL,
            estado VARCHAR(100),
            municipio VARCHAR(100),
            region_id INTEGER REFERENCES regiones(id),
            ddr VARCHAR(100),
            cader VARCHAR(100),
            ejido VARCHAR(200),
            direccion VARCHAR(500),
            localidad VARCHAR(300),
            codigo_postal VARCHAR(10),
            latitud DOUBLE PRECISION NOT NULL,
            longitud DOUBLE PRECISION NOT NULL,
            capacidad_toneladas DOUBLE PRECISION DEFAULT 0,
            cvegeo VARCHAR(10),
            cve_ent VARCHAR(5),
            cve_mun VARCHAR(5),
            fecha_actualizacion TIMESTAMP DEFAULT NOW(),
            activo BOOLEAN DEFAULT true
        );

        CREATE INDEX idx_bodegas_estado ON bodegas(estado);
        CREATE INDEX idx_bodegas_municipio ON bodegas(municipio);
        CREATE INDEX idx_bodegas_region ON bodegas(region_id);
        CREATE INDEX idx_bodegas_clave ON bodegas(clave);
        CREATE INDEX idx_bodegas_coords ON bodegas(latitud, longitud);
    """)
    conn.commit()
    print("Schema created.")

    # --- Step 1: Build regiones from DDR per state ---
    print("Building regiones...")
    regiones_map = {}  # (ddr, estado) -> id
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        estado = row[2]  # NOMBRE ESTADO
        ddr = row[3]     # NOMBRE DDR
        if estado and ddr:
            key = (ddr.strip().upper(), estado.strip().upper())
            if key not in regiones_map:
                regiones_map[key] = None

    for (ddr, estado) in sorted(regiones_map.keys()):
        cur.execute(
            "INSERT INTO regiones (nombre, estado) VALUES (%s, %s) RETURNING id",
            (ddr.title(), estado.title()),
        )
        row = cur.fetchone()
        regiones_map[(ddr, estado)] = row["id"]
    conn.commit()
    print(f"  Inserted {len(regiones_map)} regiones.")

    # --- Step 2: Insert bodegas ---
    print("Inserting bodegas...")
    count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        clave = row[0]          # CLAVE CENTRO DE ACOPIO
        estado_raw = row[2]     # NOMBRE ESTADO
        ddr = row[3]            # NOMBRE DDR
        cader = row[4]          # NOMBRE CADER
        municipio = row[5]      # NOMBRE MUNICIPIO
        ejido = row[6]          # NOMBRE EJIDO
        nombre = row[13]        # NOMBRE CENTRO DE ACOPIO
        calle = row[14]         # CALLE
        localidad = row[15]     # LOCALIDAD
        cp = row[16]            # CODIGO POSTAL
        almacenamiento = row[17]  # TOTAL ALMACENAMIENTO
        lat = row[18]           # LATITUD (decimal)
        lng = row[19]           # LONGITUD (decimal)
        cvegeo = row[20]        # CVEGEO
        cve_ent = row[21]       # CVE_ENT
        estado_col = row[22]    # Estado (title case)
        cve_mun = row[23]       # CVE_MUN
        municipio_col = row[24] # Municipio (title case)

        if lat is None or lng is None:
            continue

        # Use title-case state/municipio from cols 22/24 if available
        estado = estado_col if estado_col else (estado_raw.title() if estado_raw else None)
        muni = municipio_col if municipio_col else (municipio.title() if municipio else None)

        # Find region
        region_id = None
        if ddr and estado_raw:
            key = (ddr.strip().upper(), estado_raw.strip().upper())
            region_id = regiones_map.get(key)

        cap = parse_storage(almacenamiento)

        cur.execute(
            """INSERT INTO bodegas
               (clave, nombre, estado, municipio, region_id, ddr, cader, ejido,
                direccion, localidad, codigo_postal, latitud, longitud,
                capacidad_toneladas, cvegeo, cve_ent, cve_mun)
               VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
            (
                str(clave).strip() if clave else None,
                str(nombre).strip() if nombre else "Sin nombre",
                estado,
                muni,
                region_id,
                ddr.strip().title() if ddr else None,
                cader.strip().title() if cader else None,
                str(ejido).strip() if ejido else None,
                str(calle).strip() if calle else None,
                str(localidad).strip() if localidad else None,
                str(cp).strip() if cp else None,
                float(lat),
                float(lng),
                cap,
                str(cvegeo).strip() if cvegeo else None,
                str(cve_ent).strip() if cve_ent else None,
                str(cve_mun).strip() if cve_mun else None,
            ),
        )
        count += 1

    conn.commit()
    print(f"  Inserted {count} bodegas.")

    # --- Verify ---
    cur.execute("SELECT COUNT(*) as n FROM bodegas")
    print(f"  Verification: {cur.fetchone()['n']} bodegas in DB")
    cur.execute("SELECT COUNT(*) as n FROM regiones")
    print(f"  Verification: {cur.fetchone()['n']} regiones in DB")

    # --- Summary per state ---
    cur.execute("""
        SELECT estado, COUNT(*) as n, SUM(capacidad_toneladas) as cap
        FROM bodegas GROUP BY estado ORDER BY n DESC
    """)
    print("\nPer-state summary:")
    for r in cur.fetchall():
        print(f"  {r['estado']}: {r['n']} bodegas, {r['cap']:,.0f} ton capacity")

    cur.close()
    conn.close()
    print("\nDone!")


if __name__ == "__main__":
    main()
